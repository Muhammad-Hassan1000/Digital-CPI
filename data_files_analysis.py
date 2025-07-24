#%%
import dotenv
import os

dotenv.load_dotenv()
DATA_DIR = os.getenv("DATA_DIR")

from pathlib import Path
import pandas as pd
import re
from openpyxl import load_workbook
# import polars as pl
from datetime import datetime, timedelta

# Added by IT
def list_files():
    current_date = datetime.now().date().isoformat()
    dir_path = os.path.join(DATA_DIR, current_date)
    file_paths = [os.path.join(dir_path, f) for f in os.listdir(dir_path) if f != "data_files.csv"]
    return file_paths


def create_csv_path(current_time: str) -> str:
    current_date = datetime.strptime(current_time, "%Y-%m-%d_%H-%M").date().isoformat()
    path = f"{DATA_DIR}/{current_date}"
    if os.path.exists(path):
        return path
    else:
        os.mkdir(path)
        return path
    

def count_rows_cols_excel(file_path, sheet_name=0):
    excel_file = pd.ExcelFile(file_path, engine='openpyxl')
    wb = load_workbook(filename=file_path, read_only=True)
    if isinstance(sheet_name, int):
        sheet = wb[wb.sheetnames[sheet_name]]
    else:
        sheet = wb[sheet_name]
    
    # Get the maximum number of rows and columns
    row_count = sheet.max_row
    col_count = sheet.max_column

    return row_count - 1, col_count  # Subtracting 1 to exclude header row

def count_rows_cols_csv(file_path):
    # Counting rows
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
        row_count = sum(1 for row in file)
    
    # Counting columns
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
        first_line = file.readline()
        col_count = len(first_line.split(','))  # Adjust the delimiter if needed

    return row_count - 1, col_count  # Subtracting 1 to exclude header row



files = list_files()

files_pattern = '(naheed|alfatah|imtiaz0|imtiaz1|foodpanda|carrefour|bata|restaurants|clothing|gold|sastaticket|meds|med_tests|other_products|books|zameen|networks|men_tailoring|doctors|cars|washing|servis|Dentist_details|school_uniform|vegetables&fruits|sabzi_market|chicken_rate|metro)'

def file_reader_meta(x):
    # x = 'ALfateh2024-10-29_16-46.csv'
    print(x)
    ext = os.path.splitext(x)[1]
    if ext=='.xlsx':
        dims = count_rows_cols_excel(x)
    else:
        dims = count_rows_cols_csv(x)
    
    rows = dims[0]
    cols = dims[1]
    size = os.path.getsize(x)
    store_name = re.findall(files_pattern, x, flags=re.IGNORECASE)[0]
    date_time = re.findall('(\\d{4}-\\d{2}-\\d{2})_(\\d{2}-\\d{2})', x)
    # Added by IT
    if date_time:
        date = date_time[0][0]
        time = date_time[0][1].replace('-','')
        time = int(time)
    else:
        timestamp = os.path.getmtime(Path(x))
        formatted_ts = datetime.fromtimestamp(timestamp)
        date = formatted_ts.date().isoformat()
        time = formatted_ts.time().isoformat(timespec="minutes").replace(':','')
        time = int(time)

    df = pd.DataFrame(
        {
            'rows':[rows],
            'cols':[cols],
            'store':[store_name],
            'date': [date],
            'time': [time],
            'size':[size/1000],
            'ext':[ext],
            'link':x
        }
    )
    return df

meta = [file_reader_meta(x) for x in files]


meta = pd.concat(meta, axis = 0).reset_index(drop = True)
# Added/Modified by IT
now = datetime.now()
now = now.strftime("%Y-%m-%d_%H-%M")
csv_path = create_csv_path(now)
meta.to_csv(f"{csv_path}/data_files.csv", 
            index = False)

meta1 = meta.loc[meta.groupby(['store', 'date'])['time'].idxmax()].reset_index(drop = True)\
        .drop(['time'], axis = 1)

# meta2 = meta.loc[meta.groupby(['store', 'date'])['rows'].idxmax()].reset_index(drop = True)
# meta2.to_clipboard(index = False)


# meta2_summ = meta2.groupby(['store', 'date'], as_index=False).agg(
#     # pl.col('rows').mean()

#     {
#         'rows':'sum'
#     }
# ).pivot(
#     index = "store", columns = "date", values = "rows"
# )

# meta2_summ.drop([
#     x for x in meta2_summ.columns if x<(
#         meta2['date'].sort_values(ascending=False).drop_duplicates().head(7).min()
#     )
# ], axis = 1).\
# sort_values(
#     [meta2['date'].sort_values(ascending=False).drop_duplicates().head(1).iloc[0]],
#     ascending=True
# )

# %%
